import ipaddress
from abc import ABC, abstractmethod
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import JsonValue

from cloud_inventory.domain.models import (
    CloudResource,
    Completeness,
    DetailLevel,
    Provider,
    Relationship,
    ResourceBatch,
    ResourceType,
)
from cloud_inventory.domain.uid import build_cloud_uid
from cloud_inventory.ingest.batch import canonical_resource_sha256, finalize_batch
from cloud_inventory.ingest.parsers.base import DetectionResult, SourceMetadata

SERVER_HEADERS = {
    "name": {"Server Name", "서버 이름"},
    "id": {"Instance ID", "인스턴스 ID"},
    "status": {"Status", "상태"},
    "region": {"Region", "리전"},
    "zone": {"Zone", "존"},
    "vpc": {"VPC", "VPC 이름"},
    "subnet": {"Subnet", "Subnet 이름"},
    "private_ip": {"Private IP", "사설 IP"},
    "public_ip": {"Public IP", "공인 IP"},
}

PUBLIC_IP_HEADERS = {
    "public_ip": {"Public IP", "공인 IP"},
    "status": {"Status", "상태"},
    "applied_server": {"Applied Server", "적용 서버"},
    "private_ip": {"Private IP", "사설 IP"},
    "vpc": {"VPC", "VPC 이름"},
    "region": {"Region", "리전"},
}

LOAD_BALANCER_HEADERS = {
    "name": {"Load Balancer Name", "로드 밸런서 이름"},
    "id": {"Instance ID", "인스턴스 ID"},
    "status": {"Status", "상태"},
    "type": {"Type", "유형"},
    "network": {"Network", "네트워크"},
    "vpc": {"VPC", "VPC 이름"},
    "subnet": {"Subnet", "Subnet 이름"},
    "ip": {"IP", "IP 주소"},
    "region": {"Region", "리전"},
}

BUCKET_HEADERS = {
    "name": {"Name", "Bucket Name", "버킷 이름"},
    "region": {"Region", "리전"},
    "size": {"Size", "사용량"},
    "created_at": {"Date Created", "생성 일시"},
}

_MAX_TOTAL_UNCOMPRESSED = 500 * 1024 * 1024
_MAX_ENTRY_UNCOMPRESSED = 250 * 1024 * 1024
_MAX_EXPANSION_RATIO = 100
_SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _inspect_workbook_archive(path: Path) -> None:
    if not is_zipfile(path):
        raise ValueError("NCP workbook must be a valid ZIP workbook")

    try:
        with ZipFile(path) as archive:
            total_size = 0
            for entry in archive.infolist():
                normalized_name = entry.filename.replace("\\", "/").casefold()
                total_size += entry.file_size
                if total_size > _MAX_TOTAL_UNCOMPRESSED:
                    raise ValueError("workbook total uncompressed size exceeds 500 MB")
                if entry.file_size > _MAX_ENTRY_UNCOMPRESSED:
                    raise ValueError("workbook entry uncompressed size exceeds 250 MB")
                if entry.file_size and (
                    entry.file_size / max(entry.compress_size, 1)
                    > _MAX_EXPANSION_RATIO
                ):
                    raise ValueError("workbook ZIP expansion ratio exceeds 100")
                if normalized_name.endswith("vbaproject.bin"):
                    raise ValueError("macro-enabled workbooks are not allowed")
                if normalized_name.startswith("xl/externallinks/"):
                    raise ValueError("workbook external links are not allowed")
                if (
                    normalized_name.startswith("xl/worksheets/")
                    and normalized_name.endswith(".xml")
                ):
                    root = ElementTree.fromstring(archive.read(entry))
                    formula = root.find(f".//{{{_SPREADSHEET_NAMESPACE}}}f")
                    if formula is not None:
                        raise ValueError("workbook formula cells are not allowed")
    except BadZipFile as error:
        raise ValueError("NCP workbook must be a valid ZIP workbook") from error


def _first_non_empty_sheet(path: Path) -> list[list[object]]:
    _inspect_workbook_archive(path)
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if any(any(value is not None and value != "" for value in row) for row in rows):
                return rows
    finally:
        workbook.close()
    raise ValueError("NCP workbook has no non-empty sheet")


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _resolve_headers(
    raw_headers: list[object],
    aliases: dict[str, set[str]],
) -> dict[str, int]:
    resolved: dict[str, int] = {}
    headers = [_text(value) for value in raw_headers]
    for canonical_name, accepted_names in aliases.items():
        matching_indexes = [
            index for index, header in enumerate(headers) if header in accepted_names
        ]
        if matching_indexes:
            resolved[canonical_name] = matching_indexes[0]
    return resolved


def _row_value(
    row: list[object],
    header_map: dict[str, int],
    name: str,
) -> str:
    index = header_map.get(name)
    return _text(row[index]) if index is not None and index < len(row) else ""


def _attributes(**values: str) -> dict[str, JsonValue]:
    return {key: value for key, value in values.items() if value}


class NcpConsoleXlsxParser(ABC):
    schema_version = "1"
    source_priority = 200
    profile_id: str
    header_aliases: dict[str, set[str]]
    required_headers: frozenset[str]

    def detect(self, path: Path, metadata: SourceMetadata) -> DetectionResult:
        del metadata
        if not is_zipfile(path):
            return DetectionResult(False, 0, "content is not an XLSX ZIP workbook")
        rows = _first_non_empty_sheet(path)
        header_map = _resolve_headers(rows[0], self.header_aliases)
        matched = self.required_headers.issubset(header_map)
        return DetectionResult(
            matched=matched,
            confidence=100 if matched else 0,
            reason=f"{self.profile_id} headers matched"
            if matched
            else f"{self.profile_id} required headers are missing",
        )

    def parse(self, path: Path, metadata: SourceMetadata) -> ResourceBatch:
        if metadata.provider is not Provider.NCP:
            raise ValueError("NCP workbook requires provider ncp")

        rows = _first_non_empty_sheet(path)
        header_map = _resolve_headers(rows[0], self.header_aliases)
        if not self.required_headers.issubset(header_map):
            raise ValueError(f"{self.profile_id} required headers are missing")

        resources: list[CloudResource] = []
        zones_by_region: dict[str, set[str]] = defaultdict(set)
        regions: set[str] = set()
        for row in rows[1:]:
            if not any(_text(value) for value in row):
                continue
            region = _row_value(row, header_map, "region") or metadata.region
            if not region:
                raise ValueError("NCP workbook row region is required")
            regions.add(region)
            zone = _row_value(row, header_map, "zone")
            if zone:
                zones_by_region[region].add(zone)
            resources.extend(self._parse_resource_row(row, header_map, region, metadata))

        resources.extend(
            self._location_resources(regions, zones_by_region, metadata)
        )
        candidates_by_uid: dict[str, list[CloudResource]] = defaultdict(list)
        for resource in resources:
            candidates_by_uid[resource.uid].append(resource)
        deduplicated = [
            max(candidates, key=canonical_resource_sha256)
            for _, candidates in sorted(candidates_by_uid.items())
        ]
        return finalize_batch(
            provider=metadata.provider,
            realm=metadata.realm,
            account_id=metadata.account_id,
            observed_at=metadata.exported_at,
            resources=deduplicated,
            parser_profile=self.profile_id,
            source_priority=self.source_priority,
            detail_level=DetailLevel.SUMMARY,
        )

    def _resource(
        self,
        *,
        metadata: SourceMetadata,
        region: str,
        resource_type: ResourceType,
        external_id: str,
        name: str,
        status: str = "unknown",
        attributes: dict[str, JsonValue] | None = None,
        relationships: list[Relationship] | None = None,
    ) -> CloudResource:
        return CloudResource(
            uid=build_cloud_uid(
                metadata.provider,
                metadata.realm,
                metadata.account_id,
                region,
                resource_type,
                external_id,
            ),
            provider=metadata.provider,
            realm=metadata.realm,
            account_id=metadata.account_id,
            region=region,
            resource_type=resource_type,
            external_id=external_id,
            name=name,
            status=status or "unknown",
            attributes=attributes or {},
            relationships=relationships or [],
            observed_at=metadata.exported_at,
            completeness=Completeness.PARTIAL,
            detail_level=DetailLevel.SUMMARY,
            source_profile=self.profile_id,
            source_priority=self.source_priority,
        )

    def _location_resources(
        self,
        regions: set[str],
        zones_by_region: dict[str, set[str]],
        metadata: SourceMetadata,
    ) -> list[CloudResource]:
        resources: list[CloudResource] = []
        for region in sorted(regions):
            zone_resources = [
                self._resource(
                    metadata=metadata,
                    region=region,
                    resource_type=ResourceType.ZONE,
                    external_id=zone,
                    name=zone,
                )
                for zone in sorted(zones_by_region[region])
            ]
            resources.extend(zone_resources)
            resources.append(
                self._resource(
                    metadata=metadata,
                    region=region,
                    resource_type=ResourceType.REGION,
                    external_id=region,
                    name=region,
                    relationships=[
                        Relationship(
                            relation_type="contains",
                            target_uid=zone_resource.uid,
                        )
                        for zone_resource in zone_resources
                    ],
                )
            )
        return resources

    @abstractmethod
    def _parse_resource_row(
        self,
        row: list[object],
        header_map: dict[str, int],
        region: str,
        metadata: SourceMetadata,
    ) -> list[CloudResource]:
        """Convert one non-empty row into canonical resources."""


class NcpServerXlsxParser(NcpConsoleXlsxParser):
    profile_id = "ncp.server_list.xlsx.v1"
    header_aliases = SERVER_HEADERS
    required_headers = frozenset({"name", "id"})

    def _parse_resource_row(
        self,
        row: list[object],
        header_map: dict[str, int],
        region: str,
        metadata: SourceMetadata,
    ) -> list[CloudResource]:
        instance_id = _row_value(row, header_map, "id")
        if not instance_id:
            raise ValueError("NCP Server row identifier is required")
        name = _row_value(row, header_map, "name") or instance_id
        vm = self._resource(
            metadata=metadata,
            region=region,
            resource_type=ResourceType.VIRTUAL_MACHINE,
            external_id=instance_id,
            name=name,
            status=_row_value(row, header_map, "status"),
            attributes=_attributes(
                vpc=_row_value(row, header_map, "vpc"),
                subnet=_row_value(row, header_map, "subnet"),
            ),
        )
        resources = [vm]
        for address_type, header_name in (
            ("private", "private_ip"),
            ("public", "public_ip"),
        ):
            address = _row_value(row, header_map, header_name)
            if not address:
                continue
            ipaddress.ip_address(address)
            resources.append(
                self._resource(
                    metadata=metadata,
                    region=region,
                    resource_type=ResourceType.IP_ADDRESS,
                    external_id=f"server:{instance_id}:{address_type}:{address}",
                    name=address,
                    attributes={"address": address},
                    relationships=[
                        Relationship(relation_type="assigned_to", target_uid=vm.uid)
                    ],
                )
            )
        return resources


class NcpPublicIpXlsxParser(NcpConsoleXlsxParser):
    profile_id = "ncp.public_ip_list.xlsx.v1"
    header_aliases = PUBLIC_IP_HEADERS
    required_headers = frozenset({"public_ip", "applied_server"})

    def _parse_resource_row(
        self,
        row: list[object],
        header_map: dict[str, int],
        region: str,
        metadata: SourceMetadata,
    ) -> list[CloudResource]:
        address = _row_value(row, header_map, "public_ip")
        if not address:
            raise ValueError("NCP Public IP row identifier is required")
        ipaddress.ip_address(address)
        server_id = _row_value(row, header_map, "applied_server")
        relationships = (
            [
                Relationship(
                    relation_type="assigned_to",
                    target_uid=build_cloud_uid(
                        metadata.provider,
                        metadata.realm,
                        metadata.account_id,
                        region,
                        ResourceType.VIRTUAL_MACHINE,
                        server_id,
                    ),
                )
            ]
            if server_id
            else []
        )
        return [
            self._resource(
                metadata=metadata,
                region=region,
                resource_type=ResourceType.IP_ADDRESS,
                external_id=f"public:{address}",
                name=address,
                status=_row_value(row, header_map, "status"),
                attributes=_attributes(
                    address=address,
                    private_ip=_row_value(row, header_map, "private_ip"),
                    vpc=_row_value(row, header_map, "vpc"),
                ),
                relationships=relationships,
            )
        ]


class NcpLoadBalancerXlsxParser(NcpConsoleXlsxParser):
    profile_id = "ncp.load_balancer_list.xlsx.v1"
    header_aliases = LOAD_BALANCER_HEADERS
    required_headers = frozenset({"name", "id"})

    def _parse_resource_row(
        self,
        row: list[object],
        header_map: dict[str, int],
        region: str,
        metadata: SourceMetadata,
    ) -> list[CloudResource]:
        instance_id = _row_value(row, header_map, "id")
        if not instance_id:
            raise ValueError("NCP Load Balancer row identifier is required")
        return [
            self._resource(
                metadata=metadata,
                region=region,
                resource_type=ResourceType.LOAD_BALANCER,
                external_id=instance_id,
                name=_row_value(row, header_map, "name") or instance_id,
                status=_row_value(row, header_map, "status"),
                attributes=_attributes(
                    type=_row_value(row, header_map, "type"),
                    network=_row_value(row, header_map, "network"),
                    vpc=_row_value(row, header_map, "vpc"),
                    subnet=_row_value(row, header_map, "subnet"),
                    ip=_row_value(row, header_map, "ip"),
                ),
            )
        ]


class NcpBucketXlsxParser(NcpConsoleXlsxParser):
    profile_id = "ncp.object_storage_bucket_list.xlsx.v1"
    header_aliases = BUCKET_HEADERS
    required_headers = frozenset({"name", "created_at"})

    def _parse_resource_row(
        self,
        row: list[object],
        header_map: dict[str, int],
        region: str,
        metadata: SourceMetadata,
    ) -> list[CloudResource]:
        name = _row_value(row, header_map, "name")
        if not name:
            raise ValueError("NCP Object Storage row identifier is required")
        return [
            self._resource(
                metadata=metadata,
                region=region,
                resource_type=ResourceType.OBJECT_BUCKET,
                external_id=name,
                name=name,
                attributes=_attributes(
                    size=_row_value(row, header_map, "size"),
                    created_at=_row_value(row, header_map, "created_at"),
                ),
            )
        ]
